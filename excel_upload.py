import argparse
import logging

import powerbiclient as PBC
import openpyxl as xl


EXCEL_TYPE_TO_POWERBI_TYPE = {"n": "Double", "d": "Date", "s": "String", "b": "Boolean"}


def main():

    parser = argparse.ArgumentParser(description='List datasets from Power-BI')
    parser.add_argument('--application_id', '-i', required=True, help='application ID to sign into with')
    parser.add_argument('--application_secret', '-s', required=True, help='application secret to sign into with')
    parser.add_argument('--username', '-u', required=True, help='username to sign into with')
    parser.add_argument('-p', '--password', required=True, default=None)
    parser.add_argument('--logging-level', '-l', choices=['debug', 'info', 'error'], default='error',
                        help='desired logging level (set to error by default)')

    parser.add_argument('workbook', help='one or more excel workbooks to publish', nargs='+')

    args = parser.parse_args()

    # Set logging level based on user input
    logging_level = getattr(logging, args.logging_level.upper())
    logging.basicConfig(level=logging_level)

    # Step 1: Sign in to server.
    powerbi_auth = PBC.PowerBIAuth(args.application_id, args.application_secret, args.username, args.password)
    server = PBC.Server()

    with server.auth.acquire_token(powerbi_auth):
        # go through all Excel files
        for fname in args.workbook:
            wb = xl.load_workbook(filename=fname)
            sheet = wb.active   # select active sheet
            sheet_name = wb.sheetnames[0]
            dataset = server.datasets.get_by_name(fname)    # lookup any existing dataset by this filename
            data = []
            for row_i, row in enumerate(sheet.iter_rows()):
                if row_i == 0:              # post dataset at the first row, or delete existing rows
                    print("post dataset definition")
                    # determine column definition: get names from row 1 and types from row 2 (openpyxl is 1-based)
                    columns = [ { "name": sheet.cell(row=1,column=i+1).value,
                                  "dataType": EXCEL_TYPE_TO_POWERBI_TYPE[c.data_type] }
                                for i, c in enumerate(sheet[2]) ]
                    print(columns)
                    if dataset is not None:
                        # truncate table
                        server.datasets.delete_rows(dataset, sheet_name)
                    else:
                        # post dataset
                        dataset = server.datasets.post_dataset(fname,
                            [
                                {
                                    "name": sheet_name,
                                    "columns": columns
                                }
                            ],
                            default_retention_policy='None'
                        )
                elif row_i % 10000 == 0:    # post rows every 10000 rows
                    print("append data and post rows, cycle {0}".format(row_i))
                    data.append({ columns[i]['name']: c.value for i, c in enumerate(row) })
                    server.datasets.post_rows(dataset, sheet_name, data)
                    data = []               # reset data block
                else:                       # otherwise, just append data
                    if row_i % 1000 == 0:   # print log line once per 1000 rows
                        print("just append data, cycle {0}".format(row_i))
                    data.append({ columns[i]['name']: c.value for i, c in enumerate(row) })
            server.datasets.post_rows(dataset, sheet_name, data)


if __name__ == '__main__':
    main()
